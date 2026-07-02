"""Corpus importer: interlinear workbook -> DRAFT words.

Rules (locked — see CLAUDE.md and docs/CERTIFICATION_DESK.md):
- Every imported word is a draft. This module NEVER sets certified.
- All-or-nothing validation: if any expected column is missing or any row
  fails to parse, the exact rows are reported and NOTHING is imported.
- Idempotent: re-runs update draft rows in place, never touch certified
  rows, and report inserted / updated / unchanged / skipped-certified.
- Empty draft glosses are legal in the source (e.g. suffix-only words the
  Strong's pipeline couldn't gloss); they import as empty drafts and the
  desk refuses to certify them until an editor writes the gloss.
"""

import openpyxl

from app import db
from app.models import Track, Unit, Word

LADDER = [
    ("chumash", "Chumash — Noach & Lech Lecha", 1),
    ("avos", "Pirkei Avos", 2),
    ("eilu_metzios", "Eilu Metzios (Bava Metzia 21a–33b)", 3),
]

# Standard aliyah boundaries for Parshas Noach (Bereishis 6:9–11:32),
# inclusive (chapter, pasuk) ranges.
NOACH_ALIYOS = [
    ((6, 9), (6, 22), 1),
    ((7, 1), (7, 16), 2),
    ((7, 17), (8, 14), 3),
    ((8, 15), (9, 7), 4),
    ((9, 8), (9, 17), 5),
    ((9, 18), (10, 32), 6),
    ((11, 1), (11, 32), 7),
]

SHEET = "Interlinear"
REQUIRED_COLUMNS = {
    "Reference": "ref",
    "Ch": "ch",
    "V": "v",
    "Word #": "word_n",
    "Hebrew (WLC)": "hebrew",
    "Shoresh": "shoresh",
}
GLOSS_COLUMN_PREFIX = "Literal English"


class ImportValidationError(Exception):
    def __init__(self, problems):
        self.problems = problems
        super().__init__(
            f"{len(problems)} validation problem(s); nothing imported:\n"
            + "\n".join(f"  - {p}" for p in problems[:40])
            + ("" if len(problems) <= 40 else f"\n  ... and {len(problems) - 40} more")
        )


def aliyah_for(ch, v):
    for lo, hi, n in NOACH_ALIYOS:
        if lo <= (ch, v) <= hi:
            return n
    return None


def seed_ladder():
    for key, name, rung in LADDER:
        if not Track.query.filter_by(key=key).first():
            db.session.add(Track(key=key, name=name, rung_order=rung))
    db.session.commit()


def parse_interlinear(path):
    """Read and validate the workbook. Returns (word_dicts, empty_gloss_rows)
    or raises ImportValidationError without touching the database."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ImportValidationError(
            [f"sheet {SHEET!r} not found (sheets: {wb.sheetnames})"])
    rows = list(wb[SHEET].iter_rows(values_only=True))
    if not rows:
        raise ImportValidationError([f"sheet {SHEET!r} is empty"])

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {}
    for name, key in REQUIRED_COLUMNS.items():
        if name not in header:
            col[key] = None
        else:
            col[key] = header.index(name)
    gloss_idx = next((i for i, h in enumerate(header)
                      if h.startswith(GLOSS_COLUMN_PREFIX)), None)
    missing = [name for name, key in REQUIRED_COLUMNS.items()
               if col[key] is None]
    if gloss_idx is None:
        missing.append(f"{GLOSS_COLUMN_PREFIX}… (gloss)")
    if missing:
        raise ImportValidationError(
            [f"missing expected column(s): {', '.join(missing)}"])

    def cell(r, key):
        return r[col[key]] if col[key] < len(r) else None

    problems, words, empty_gloss_rows = [], [], []
    pasuk_index_of = {}
    position = 0
    for rownum, r in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in r):
            continue  # ignore fully blank trailing rows
        ref = cell(r, "ref")
        if not (isinstance(ref, str) and ref.strip()):
            problems.append(f"row {rownum}: empty or non-text Reference ({ref!r})")
            continue
        try:
            ch, v, word_n = (int(cell(r, "ch")), int(cell(r, "v")),
                             int(cell(r, "word_n")))
        except (TypeError, ValueError):
            problems.append(
                f"row {rownum} ({ref}): non-numeric Ch/V/Word # "
                f"({cell(r, 'ch')!r}, {cell(r, 'v')!r}, {cell(r, 'word_n')!r})")
            continue
        hebrew = cell(r, "hebrew")
        if not (isinstance(hebrew, str) and hebrew.strip()):
            problems.append(f"row {rownum} ({ref}): empty Hebrew (WLC)")
            continue
        gloss = r[gloss_idx] if gloss_idx < len(r) else None
        gloss = str(gloss).strip() if gloss is not None else ""
        if not gloss:
            empty_gloss_rows.append((rownum, ref.strip(), word_n))
        shoresh = cell(r, "shoresh")
        shoresh = str(shoresh).strip() if shoresh else None

        key = (ch, v)
        if key not in pasuk_index_of:
            pasuk_index_of[key] = len(pasuk_index_of) + 1
        position += 1
        words.append({
            "ref": ref.strip(),
            "pasuk_index": pasuk_index_of[key],
            "position": position,
            "hebrew": hebrew.strip(),
            "translation": gloss,
            "shoresh": shoresh,
            "aliyah": aliyah_for(ch, v),
        })

    if problems:
        raise ImportValidationError(problems)
    return words, empty_gloss_rows


def import_words(words, track_key="chumash", unit_name="Noach", kind="parsha"):
    """Idempotent upsert into one unit. Never touches certified rows and
    never writes the certified flag."""
    track = Track.query.filter_by(key=track_key).one()
    unit = Unit.query.filter_by(track_id=track.id, name=unit_name).first()
    if unit is None:
        unit = Unit(track_id=track.id, name=unit_name, kind=kind,
                    order_index=Unit.query.filter_by(track_id=track.id).count() + 1)
        db.session.add(unit)
        db.session.flush()

    existing = {w.position: w for w in Word.query.filter_by(unit_id=unit.id)}
    counts = {"inserted": 0, "updated": 0, "unchanged": 0, "skipped_certified": 0}
    fields = ("ref", "pasuk_index", "hebrew", "translation", "shoresh", "aliyah")

    for d in words:
        row = existing.get(d["position"])
        if row is None:
            db.session.add(Word(unit_id=unit.id, **d))
            counts["inserted"] += 1
        elif row.certified:
            counts["skipped_certified"] += 1
        elif any(getattr(row, f) != d[f] for f in fields):
            for f in fields:
                setattr(row, f, d[f])
            counts["updated"] += 1
        else:
            counts["unchanged"] += 1

    db.session.commit()
    return unit, counts
